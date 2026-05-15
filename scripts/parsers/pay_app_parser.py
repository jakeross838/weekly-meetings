"""Parser for Ross Built AIA G702/G703 pay app workbooks.

Public entrypoint:
    parse_pay_app(file_path: Path, job_id: str) -> dict

Returns:
    {
        "pay_app":      dict mapping pay_apps table columns to values,
        "line_items":   list of dicts mapping pay_app_line_items columns,
        "skipped_rows": list of {row, reason, ...} for debugging only,
    }

Sheet handling:
- Prefers exact sheet names "Project Summary (G702)" and "Line Item Estimate (G703)".
- Falls back to a substring match on "G702" / "G703" if exact is missing.

G703 column mapping is done by header text (the two-row header block above
the data rows is combined and matched against known label patterns). This
keeps the parser robust to column-letter drift across jobs.

No Supabase I/O. Callers are responsible for ingestion.
"""

from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


SHEET_G702_EXACT = "Project Summary (G702)"
SHEET_G703_EXACT = "Line Item Estimate (G703)"

_EXCEL_ERRORS = {
    "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#GETTING_DATA",
}


def _to_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s in _EXCEL_ERRORS:
            return None
        s = s.replace("$", "").replace(",", "").strip()
        if s in ("", "-"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _to_text(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v)


def _normalize(s: str) -> str:
    """Uppercase, strip dot-fill padding, collapse whitespace."""
    s = s.upper()
    s = re.sub(r"\.+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _find_sheet(wb: Workbook, exact: str, fuzzy_substring: str) -> Worksheet:
    if exact in wb.sheetnames:
        return wb[exact]
    matches = [n for n in wb.sheetnames if fuzzy_substring.lower() in n.lower()]
    if matches:
        return wb[matches[0]]
    raise ValueError(
        f"No sheet matching exact={exact!r} or substring={fuzzy_substring!r}; "
        f"available sheets: {wb.sheetnames}"
    )


def _sheet_to_json(ws: Worksheet) -> dict:
    """Serialize every non-null cell to {cell_address: value}.
    Dates become ISO strings so the result is JSON-safe."""
    out: dict[str, Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, datetime):
                out[cell.coordinate] = v.isoformat()
            elif isinstance(v, date):
                out[cell.coordinate] = v.isoformat()
            elif isinstance(v, (int, float, str, bool)):
                out[cell.coordinate] = v
            else:
                out[cell.coordinate] = str(v)
    return out


def _find_label_value(ws: Worksheet, patterns: list[str]) -> Any:
    """Scan the sheet for a cell whose normalized text contains any pattern,
    then scan rightward on that row for the first non-decorator value.
    Returns the raw cell value or None."""
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            norm = _normalize(cell.value)
            if not any(p in norm for p in patterns):
                continue
            for c in range(cell.column + 1, ws.max_column + 1):
                v = ws.cell(row=cell.row, column=c).value
                if v is None:
                    continue
                if isinstance(v, (datetime, date)):
                    return v
                if isinstance(v, bool):
                    continue
                if isinstance(v, (int, float)):
                    return v
                if isinstance(v, str):
                    stripped = v.strip()
                    if stripped in ("$", "-", "", ":"):
                        continue
                    n = _to_number(v)
                    if n is not None:
                        return n
                    return v
    return None


G703_COLUMN_PATTERNS: list[tuple[str, list[str]]] = [
    ("line_number",                 ["ITEM NO", "ITEM"]),
    ("description",                 ["DESCRIPTION OF WORK", "DESCRIPTION"]),
    ("scheduled_value",             ["SCHEDULED VALUE", "ORIGINAL ESTIMATE", "ORIGINAL"]),
    ("work_completed_previous",     ["PREVIOUS APPLICATIONS", "WORK COMPLETED FROM PREVIOUS", "FROM PREVIOUS", "PREVIOUS"]),
    ("work_completed_this_period",  ["REQUISITION THIS PERIOD", "WORK COMPLETED THIS PERIOD", "THIS PERIOD"]),
    ("materials_stored",            ["MATERIALS PRESENTLY STORED", "MATERIALS STORED"]),
    ("total_completed",             ["TOTAL COMPLETED AND STORED", "TOTAL TO DATE", "TOTAL COMPLETED"]),
    ("pct_complete",                ["% COMP", "PERCENT COMPLETE", "PCT COMPLETE", "%"]),
    ("balance_to_finish",           ["BALANCE TO FINISH"]),
    ("retainage",                   ["RETAINAGE"]),
]


def _find_header_row(ws: Worksheet, max_scan: int = 30) -> int:
    """Return the index of the first row containing 'DESCRIPTION OF WORK'."""
    for r in range(1, max_scan + 1):
        row_text_parts = []
        for c in range(1, min(ws.max_column, 14) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_text_parts.append(_normalize(str(v)))
        row_text = " ".join(row_text_parts)
        if "DESCRIPTION OF WORK" in row_text and ("ITEM" in row_text or "NO" in row_text):
            return r
    raise ValueError("Could not locate G703 header row")


def _build_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Combine header text from header_row and header_row+1 for each column,
    then map schema fields to column indexes by pattern match."""
    combined: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        parts = []
        for r in (header_row, header_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                parts.append(_normalize(str(v)))
        if parts:
            combined[c] = " ".join(parts).strip()

    col_map: dict[str, int] = {}
    used_cols: set[int] = set()
    for field, patterns in G703_COLUMN_PATTERNS:
        for c, header in combined.items():
            if c in used_cols:
                continue
            if any(p in header for p in patterns):
                col_map[field] = c
                used_cols.add(c)
                break
    return col_map


def _derive_division(line_number: str) -> str | None:
    """CSI-style division from a numeric prefix (e.g. '01' from '01105').
    Returns None when the line number doesn't begin with two digits."""
    m = re.match(r"^(\d{2})\d", line_number.strip())
    return m.group(1) if m else None


def _parse_g703(ws: Worksheet) -> tuple[list[dict], list[dict]]:
    header_row = _find_header_row(ws)
    col_map = _build_column_map(ws, header_row)

    required = ("line_number", "description", "scheduled_value")
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(f"G703 missing required columns: {missing}. Detected: {col_map}")

    data_start = header_row + 2
    candidates: list[dict] = []
    skipped: list[dict] = []
    pct_values: list[float] = []

    for r in range(data_start, ws.max_row + 1):
        ln_raw = ws.cell(row=r, column=col_map["line_number"]).value
        desc_raw = ws.cell(row=r, column=col_map["description"]).value
        sv_raw = ws.cell(row=r, column=col_map["scheduled_value"]).value

        if ln_raw is None or (isinstance(ln_raw, str) and not ln_raw.strip()):
            if desc_raw is not None or sv_raw is not None:
                skipped.append({
                    "row": r,
                    "reason": "missing line_number",
                    "description": _to_text(desc_raw),
                })
            continue

        line_number = str(ln_raw).strip()
        description = _to_text(desc_raw)
        scheduled_value = _to_number(sv_raw)

        if description is None:
            skipped.append({"row": r, "line_number": line_number, "reason": "missing description"})
            continue
        if scheduled_value is None:
            skipped.append({
                "row": r,
                "line_number": line_number,
                "description": description,
                "reason": "missing scheduled_value",
            })
            continue

        norm_desc = _normalize(description)
        if norm_desc.startswith("TOTAL") or norm_desc in {"TOTALS", "GRAND TOTAL", "SUBTOTAL"}:
            skipped.append({
                "row": r,
                "line_number": line_number,
                "description": description,
                "reason": "looks like a total/subtotal row",
            })
            continue

        def _opt(field: str) -> float | None:
            return _to_number(ws.cell(row=r, column=col_map[field]).value) if field in col_map else None

        item = {
            "line_number":                line_number,
            "description":                description,
            "division":                   _derive_division(line_number),
            "scheduled_value":            scheduled_value,
            "work_completed_previous":    _opt("work_completed_previous"),
            "work_completed_this_period": _opt("work_completed_this_period"),
            "materials_stored":           _opt("materials_stored"),
            "total_completed":            _opt("total_completed"),
            "pct_complete":               _opt("pct_complete"),
            "balance_to_finish":          _opt("balance_to_finish"),
            "retainage":                  _opt("retainage"),
            "raw_row_index":              r,
        }
        if item["pct_complete"] is not None:
            pct_values.append(item["pct_complete"])
        candidates.append(item)

    # Detect 0-100 vs 0-1 storage format. Krauss has real overshoots up to
    # ~180%, so threshold has to sit well above 2.0. A 0-100 column has values
    # >10 in normal operation; >5 here is a safe discriminator.
    if pct_values and max(pct_values) > 5:
        for item in candidates:
            if item["pct_complete"] is not None:
                item["pct_complete"] = item["pct_complete"] / 100.0

    return candidates, skipped


def parse_pay_app(file_path: Path | str, job_id: str) -> dict:
    """Parse a Ross Built AIA G702/G703 pay app workbook.

    See module docstring for return shape. No DB I/O.
    """
    file_path = Path(file_path)
    file_hash = _sha256_file(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    ws_g702 = _find_sheet(wb, SHEET_G702_EXACT, "G702")
    ws_g703 = _find_sheet(wb, SHEET_G703_EXACT, "G703")

    raw_pa_num = _find_label_value(ws_g702, ["APPLICATION NO", "APPLICATION NUMBER"])
    pa_num_n = _to_number(raw_pa_num)
    pay_app_number = int(pa_num_n) if pa_num_n is not None else None

    pay_app = {
        "job_id":                 job_id,
        "pay_app_number":         pay_app_number,
        "application_date":       _to_date(_find_label_value(ws_g702, ["APPLICATION DATE"])),
        "contract_amount":        _to_number(_find_label_value(ws_g702, ["ORIGINAL CONTRACT SUM"])),
        "total_completed_stored": _to_number(_find_label_value(
            ws_g702, ["TOTAL COMPLETED & STORED", "TOTAL COMPLETED AND STORED"])),
        "retainage":              _to_number(_find_label_value(ws_g702, ["LESS RETAINAGE", "RETAINAGE"])),
        "current_payment_due":    _to_number(_find_label_value(ws_g702, ["CURRENT PAYMENT DUE"])),
        "source_file_name":       file_path.name,
        "source_file_hash":       file_hash,
        "raw_g702_json":          _sheet_to_json(ws_g702),
    }

    line_items, skipped = _parse_g703(ws_g703)
    for li in line_items:
        li["job_id"] = job_id

    return {"pay_app": pay_app, "line_items": line_items, "skipped_rows": skipped}
